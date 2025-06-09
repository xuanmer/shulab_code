#!/usr/bin/env python3
"""
Batch Extraction and Organization of FreeSurfer Statistics

Author: bty
Date: 2025-06-09

Description:
    This script automates the extraction of surface and subcortical statistics from
    multiple FreeSurfer-processed subjects using various parcellations. It generates
    summary CSVs and Excel files for downstream analysis.

Directory Structure Example:
tmp/
├── code/
│   └── run_stats_batch.py    # ← This script
├── data/
│   ├── 1000037/
│   │   └── FreeSurfer/
│   │       └── ... (stats/, label/, etc.)
│   ├── 1000043/
│   │   └── FreeSurfer/
│   │       └── ...
│   └── ...
└── ...

Usage:
    python run_stats_batch.py
    - All logs will be written to process_stats.log under code/
"""

import os
import subprocess
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor 

# ====== Configurations ======
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CODE_DIR = os.path.join(PROJECT_ROOT, "code")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
LOG_PATH = os.path.join(CODE_DIR, "process_stats.log")

SUBJ_PATH_FILE = os.path.join(CODE_DIR, "subj_path.txt")
GLOBAL_CSV = os.path.join(CODE_DIR, "aparc", "Global.csv")

HEMISPHERES = ["lh", "rh"]
MEASUREMENTS = ["area", "volume", "thickness", "meancurv", "gauscurv", "foldind", "curvind"]
PARCELLATIONS = ["aparc", "BN_Atlas", "Schaefer200", "hcp-mmp-b"]

SUBCORTICAL_STATS = [
    {
        "name": "aseg",
        "file": "aseg.stats",
        "output_dir": "aparc",
        "excel_name": "aseg_stats.xlsx",
        "sheet_name": "aseg.table"
    },
    {
        "name": "BN_Atlas_subcotex",
        "file": "BN_Atlas_subcotex.stats",
        "output_dir": "BN_Atlas",
        "excel_name": "BN_Atlas_subcotex_stats.xlsx",
        "sheet_name": "stats"
    }
]

def write_log(msg):
    """
    Write a log message to LOG_PATH with a timestamp.

    Parameters:
        msg (str): The log message to be written.

    Returns:
        None

    Side effects:
        Appends a line to LOG_PATH.
    """
    with open(LOG_PATH, "a") as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {msg}\n")

def generate_subject_path_file():
    """
    Scan the data directory for subjects, and create a text file with a line for each
    subject's FreeSurfer output directory.

    Returns:
        int: Number of subjects found and written.

    Side effects:
        - Creates/overwrites the file SUBJ_PATH_FILE.
        - Prints information and writes a log.

    Typical errors:
        - DATA_DIR missing or unreadable: will raise OSError.
    """
    os.makedirs(CODE_DIR, exist_ok=True)
    subject_ids = [d for d in os.listdir(DATA_DIR) if os.path.isdir(os.path.join(DATA_DIR, d))]
    with open(SUBJ_PATH_FILE, 'w') as f:
        for subj_id in subject_ids:
            f.write(f"{DATA_DIR}/{subj_id}/FreeSurfer\n")
    print(f"[INFO] Generated subject path file: {SUBJ_PATH_FILE}")
    write_log(f"Generated subject path file for {len(subject_ids)} subjects.")
    return len(subject_ids)

def create_parc_directories():
    """
    Create output directories for each parcellation under CODE_DIR.

    Returns:
        dict: A mapping from parcellation name to its absolute output directory path.

    Side effects:
        - May create new directories.

    Typical errors:
        - Insufficient permissions to create directories: raises OSError.
    """
    parc_dirs = {}
    for parc in PARCELLATIONS:
        parc_dir = os.path.join(CODE_DIR, parc)
        os.makedirs(parc_dir, exist_ok=True)
        parc_dirs[parc] = parc_dir
        print(f"[INFO] Created parcellation output directory: {parc_dir}")
    return parc_dirs

def run_one(task):
    """
    Thread worker: print command info, then execute command, print and log result.
    task: (cmd, output_file, idx, total)
    """
    cmd, output_file, idx, total = task
    cmd_str = ' '.join(cmd)
    print(f"[INFO] ({idx}/{total}) Running: {cmd_str}")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            print(f"    [OK] Output: {output_file}")
            write_log(f"aparcstats2table success: {output_file}")
        else:
            print(f"    [FAIL] Output: {output_file}\n    Error: {result.stderr or result.stdout}")
            write_log(f"aparcstats2table FAIL: {output_file} -- {result.stderr or result.stdout}")
    except Exception as e:
        print(f"[EXCEPTION] Output: {output_file}\n    Exception: {str(e)}")
        write_log(f"aparcstats2table EXCEPTION: {output_file} -- {str(e)}")

def run_aparcstats2table_parallel(subj_count, parc_dirs, max_workers=8):
    """
    Batch run aparcstats2table across all combinations of hemisphere, measurement,
    and parcellation, using concurrent processing for speed.

    Parameters:MAX_THREADS 
        subj_count (int): Number of subjects to process.
        parc_dirs (dict): Mapping of parcellation names to their output dirs.
        max_workers (int): Number of worker processes.

    Returns:
        None

    Side effects:
        - Runs up to max_workers processes in parallel.
        - Creates output CSVs in the respective directories.
        - Prints/logs status.

    Typical errors:
        - If subj_count == 0, returns immediately.
        - Subprocess or filesystem errors may print/log failures.
    """
    if subj_count == 0:
        print("[WARN] No subjects found.")
        return

    tasks = []
    total = len(HEMISPHERES) * len(MEASUREMENTS) * len(PARCELLATIONS)
    idx = 1
    for hemi in HEMISPHERES:
        for meas in MEASUREMENTS:
            for parc in PARCELLATIONS:
                parc_output_dir = parc_dirs[parc]
                output_file = os.path.join(parc_output_dir, f"{hemi}.{parc}.{meas}.csv")
                cmd = [
                    "aparcstats2table",
                    f"--subjectsfile={SUBJ_PATH_FILE}",
                    f"--hemi={hemi}",
                    f"--meas={meas}",
                    f"--parc={parc}",
                    f"--tablefile={output_file}",
                    "--delimiter=comma",
                    "--skip"
                ]
                tasks.append((cmd, output_file, idx, total))
                idx += 1

    print(f"[INFO] Running {total} aparcstats2table commands in threads...")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        list(executor.map(run_one, tasks))

def run_asegstats2table(subj_count, parc_dirs):
    """
    Run asegstats2table to extract subcortical volumes for each stat type in SUBCORTICAL_STATS.

    Parameters:
        subj_count (int): Number of subjects to process.
        parc_dirs (dict): Mapping of parcellation names to their output dirs.

    Returns:
        None

    Side effects:
        - Writes CSV files to output directories.
        - Prints and logs results.

    Typical errors:
        - Subprocess failures or missing files print/log errors.
    """
    if subj_count == 0:
        print("[WARN] No subjects found.")
        return
    total_commands = len(SUBCORTICAL_STATS)
    print(f"[INFO] Running {total_commands} asegstats2table commands...")
    command_counter = 0
    for stats_info in SUBCORTICAL_STATS:
        stats_name = stats_info["name"]
        stats_file = stats_info["file"]
        output_dir_name = stats_info["output_dir"]
        output_dir = parc_dirs[output_dir_name]
        command_counter += 1
        output_file = os.path.join(output_dir, f"{stats_name}.volume.csv")
        cmd = [
            "asegstats2table",
            f"--subjectsfile={SUBJ_PATH_FILE}",
            "--meas=volume",
            f"--statsfile={stats_file}",
            f"--tablefile={output_file}",
            "--delimiter=comma",
            "--skip"
        ]
        print(f"[INFO] ({command_counter}/{total_commands}) Running: {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            print(f"    [OK] Output: {output_file}")
            write_log(f"asegstats2table success: {output_file}")
        else:
            error_msg = result.stderr.strip() or result.stdout.strip()
            print(f"    [FAIL] Return code: {result.returncode}\n    Error: {error_msg}")
            write_log(f"asegstats2table FAIL: {output_file} -- {error_msg}")

def extract_global_metrics(subject_dir):
    """
    Extract global-level metrics (GMV, sGMV, WMV, Ventricles, TCV, vertex counts,
    mean thickness, total area) for a single subject from FreeSurfer stats files.

    Parameters:
        subject_dir (str): Path to the subject's FreeSurfer output directory.

    Returns:
        dict: {metric_name: value, ...} for the subject.

    Typical errors:
        - Missing stats files: all metrics set to NaN, warning printed.
        - Regexp pattern mismatch: metric set to NaN, warning printed.
    """
    subj_id = os.path.basename(os.path.dirname(subject_dir))
    metrics = {"case_dir": subj_id}
    metrics_rules = {
        "aseg": [
            ("GMV", r"Measure Cortex, CortexVol, Total cortical gray matter volume, (\S+), mm\^3"),
            ("sGMV", r"Measure SubCortGray, SubCortGrayVol, Subcortical gray matter volume, (\S+), mm\^3"),
            ("WMV", r"Measure CerebralWhiteMatter, CerebralWhiteMatterVol, Total cerebral white matter volume, (\S+), mm\^3"),
            ("Ventricles", r"Measure VentricleChoroidVol, VentricleChoroidVol, Volume of ventricles and choroid plexus, (\S+), mm\^3"),
            ("TCV", r"Measure EstimatedTotalIntraCranialVol, eTIV, Estimated Total Intracranial Volume, (\S+), mm\^3")
        ],
        "aparc": [
            ("lhVertex", r"Measure Cortex, NumVert, Number of Vertices, (\d+), unitless"),
            ("lh_totaISA2", r"Measure Cortex, WhiteSurfArea, White Surface Total Area, (\S+), mm\^2"),
            ("lhMeanThickness", r"Measure Cortex, MeanThickness, Mean Thickness, (\S+), mm")
        ]
    }
    # Extract from aseg.stats
    aseg_stats = os.path.join(subject_dir, "stats", "aseg.stats")
    if os.path.exists(aseg_stats):
        with open(aseg_stats, 'r') as f:
            content = f.read()
            for metric, pattern in metrics_rules["aseg"]:
                match = re.search(pattern, content)
                if match:
                    metrics[metric] = float(match.group(1))
                else:
                    print(f"[WARN] {subj_id} missing {metric} in aseg.stats.")
                    metrics[metric] = float('nan')
    # Extract from lh.aparc.stats
    lh_stats = os.path.join(subject_dir, "stats", "lh.aparc.stats")
    if os.path.exists(lh_stats):
        with open(lh_stats, 'r') as f:
            content = f.read()
            for metric, pattern in metrics_rules["aparc"]:
                match = re.search(pattern, content)
                if match:
                    value = match.group(1)
                    metrics[metric] = int(value) if metric == "lhVertex" else float(value)
                else:
                    print(f"[WARN] {subj_id} missing {metric} in lh.aparc.stats.")
                    metrics[metric] = float('nan')
    # Extract from rh.aparc.stats
    rh_stats = os.path.join(subject_dir, "stats", "rh.aparc.stats")
    if os.path.exists(rh_stats):
        with open(rh_stats, 'r') as f:
            content = f.read()
            for metric, pattern in metrics_rules["aparc"]:
                rh_metric = metric.replace("lh", "rh")
                rh_pattern = pattern.replace("lh", "rh")
                match = re.search(rh_pattern, content)
                if match:
                    value = match.group(1)
                    metrics[rh_metric] = int(value) if rh_metric == "rhVertex" else float(value)
                else:
                    print(f"[WARN] {subj_id} missing {rh_metric} in rh.aparc.stats.")
                    metrics[rh_metric] = float('nan')
    # Derived metrics: meanCT2 = mean cortical thickness (weighted by vertex count)
    try:
        if all(k in metrics for k in ["lhVertex", "lhMeanThickness", "rhVertex", "rhMeanThickness"]):
            total_vertices = metrics["lhVertex"] + metrics["rhVertex"]
            if total_vertices > 0:
                metrics["meanCT2"] = (
                    metrics["lhVertex"] * metrics["lhMeanThickness"] +
                    metrics["rhVertex"] * metrics["rhMeanThickness"]
                ) / total_vertices
            else:
                metrics["meanCT2"] = float('nan')
                print(f"[WARN] {subj_id} vertex count is zero for meanCT2.")
        else:
            metrics["meanCT2"] = float('nan')
            print(f"[WARN] {subj_id} missing vertex/thickness info for meanCT2.")
    except Exception as e:
        metrics["meanCT2"] = float('nan')
        print(f"[WARN] {subj_id} error computing meanCT2: {str(e)}")
    # Derived metrics: totalSA2 = sum of lh and rh total area
    try:
        if all(k in metrics for k in ["lh_totaISA2", "rh_totaISA2"]):
            metrics["totalSA2"] = metrics["lh_totaISA2"] + metrics["rh_totaISA2"]
        else:
            metrics["totalSA2"] = float('nan')
            print(f"[WARN] {subj_id} missing total area info for totalSA2.")
    except Exception as e:
        metrics["totalSA2"] = float('nan')
        print(f"[WARN] {subj_id} error computing totalSA2: {str(e)}")

    return metrics

def generate_global_csv():
    """
    Aggregate global metrics for all subjects and save as a CSV file for downstream analysis.

    Returns:
        None

    Side effects:
        - Creates/overwrites GLOBAL_CSV file.
        - Prints progress and warnings.
    """
    print("[INFO] Generating Global.csv ...")
    os.makedirs(os.path.dirname(GLOBAL_CSV), exist_ok=True)
    subject_ids = [d for d in os.listdir(DATA_DIR) if os.path.isdir(os.path.join(DATA_DIR, d))]
    all_metrics = []
    for i, subj_id in enumerate(subject_ids, 1):
        print(f"[INFO] Processing {i}/{len(subject_ids)}: {subj_id}")
        subject_dir = os.path.join(DATA_DIR, subj_id, "FreeSurfer")
        try:
            metrics = extract_global_metrics(subject_dir)
            all_metrics.append(metrics)
        except Exception as e:
            print(f"[FAIL] {subj_id} extraction error: {str(e)}")
    if not all_metrics:
        print("[FAIL] No global metrics extracted; aborting CSV generation.")
        return
    columns_order = [
        "case_dir", "GMV", "sGMV", "WMV", "Ventricles",
        "lhVertex", "rhVertex", "lhMeanThickness", "rhMeanThickness",
        "meanCT2", "lh_totaISA2", "rh_totaISA2", "totalSA2", "TCV"
    ]
    df = pd.DataFrame(all_metrics)
    for col in columns_order:
        if col not in df.columns:
            df[col] = float('nan')
    df = df[columns_order]
    df.to_csv(GLOBAL_CSV, index=False)
    print(f"[OK] Global.csv saved: {GLOBAL_CSV}")

def merge_parc_to_excel(parc):
    """
    For a given parcellation, merge all corresponding CSV files in its output directory
    into a multi-sheet Excel file. The first sheet for 'aparc' will be Global.csv.

    Parameters:
        parc (str): Parcellation name.

    Returns:
        None

    Side effects:
        - Writes an Excel file with multiple sheets.
        - Prints/logs status.

    Typical errors:
        - File not found, openpyxl errors, or sheet name too long.
    """
    parc_dir = os.path.join(CODE_DIR, parc)
    if not os.path.exists(parc_dir):
        print(f"[FAIL] Parcellation directory missing: {parc_dir}")
        return
    excel_file = os.path.join(CODE_DIR, f"{parc}_stats.xlsx")
    wb = Workbook()
    default_sheet_name = "Sheet"
    # Sheet: Global.csv (first, only for 'aparc')
    if parc == "aparc":
        file_path = os.path.join(parc_dir, "Global.csv")
        sheet_name = "Global.table"
        if os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)
                ws = wb.create_sheet(title=sheet_name, index=0)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                print(f"[OK] Added Global.csv to Excel: {sheet_name}")
            except Exception as e:
                print(f"[FAIL] Global.csv failed: {str(e)}")
    # Other CSVs
    for file in os.listdir(parc_dir):
        if file.endswith('.csv') and file != 'Global.csv':
            file_path = os.path.join(parc_dir, file)
            sheet_name = file.replace('.csv', '.table')
            try:
                df = pd.read_csv(file_path)
                if not df.empty:
                    df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: os.path.basename(os.path.dirname(x)))
                ws = wb.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                print(f"[OK] Added: {file_path} → {sheet_name}")
            except Exception as e:
                print(f"[FAIL] {file_path} failed: {str(e)}")
    if default_sheet_name in wb.sheetnames:
        del wb[default_sheet_name]
    wb.save(excel_file)
    print(f"[OK] Excel saved: {excel_file}")

def generate_all_excel_files():
    """
    For all parcellations, merge their CSVs into Excel summary files.

    Returns:
        None

    Side effects:
        - Calls merge_parc_to_excel for each parcellation.
        - Prints/logs status.
    """
    print("[INFO] Generating Excel files for all parcellations ...")
    for parc in PARCELLATIONS:
        print(f"[INFO] Merging CSVs for: {parc}")
        merge_parc_to_excel(parc)
    print("[OK] All Excel files generated.")

def main():
    """
    Main entry point. Orchestrates the full workflow of:
      1. Generating the subject list.
      2. Creating output directories.
      3. Running parallel aparcstats2table extraction.
      4. Running subcortical asegstats2table extraction.
      5. Summarizing global metrics.
      6. Generating Excel files.

    Returns:
        None

    Side effects:
        - Triggers all the file IO and subprocess work.
        - Prints/logs workflow steps.
    """
    print("===== Batch FreeSurfer Statistics Processing START =====")
    write_log("=== BATCH START ===")
    subj_count = generate_subject_path_file()
    parc_dirs = create_parc_directories()
    run_aparcstats2table_parallel(subj_count, parc_dirs, max_workers=8)
    run_asegstats2table(subj_count, parc_dirs)
    generate_global_csv()
    generate_all_excel_files()
    write_log("=== BATCH FINISHED ===")
    print("===== All Processing Done =====")

if __name__ == "__main__":
    main()
